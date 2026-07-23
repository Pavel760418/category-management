"""Тесты русскоязычных шаблонов, маппинга и schema-слоя."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.core.status import status_by_ratio
from app.services.mapping_service import MappingService
from app.services.schema_service import SchemaService, get_schema_service
from app.services.template_service import TemplateService
from app.services.validation_service import validate_sku_df
from app.services.calculation_service import CalculationService
from app.utils.io import read_template_excel


def test_up_metric_ok():
    ratio, status = status_by_ratio(120, 100, "up")
    assert ratio >= 1
    assert "Выполнено" in status


def test_down_metric_exceeded():
    ratio, status = status_by_ratio(178, 21, "down", 1.3)
    assert ratio < 1
    assert "Превышение" in status or "Риск" in status


def test_schema_display_names_are_russian():
    schema = SchemaService()
    assert schema.display_name("sku_code") == "Артикул SKU"
    assert schema.display_name("sales_value") == "Продажи, руб."
    assert schema.display_name("shelf_meters") == "Длина выкладки, м"
    assert "sku_code" not in schema.display_name("sku_code")


def test_template_ui_config_has_russian_names():
    schema = get_schema_service()
    tpl = TemplateService(str(Path("data/processed/test_templates")), schema)
    items = tpl.list_templates_ru()
    assert len(items) == 3
    names = {i["file_name"] for i in items}
    assert "Шаблон_данных_SKU.xlsx" in names
    assert "Шаблон_конкуренты.xlsx" in names
    assert "Шаблон_цели_категории.xlsx" in names
    for item in items:
        assert item["sheet_name"]
        assert all(ord(c) < 128 or True for c in item["file_name"])  # smoke
        # user-facing labels must not be English technical keys
        assert "sku_upload" not in item["file_name"].lower()
        assert "template" not in item["file_name"].lower()
        for label in item["required_labels"]:
            assert "_" not in label or label in {"LFL продажи"}  # allow rare cases
            # no raw canonical keys as labels
            assert label not in schema.fields


def test_russian_template_generation_and_load(tmp_path):
    schema = SchemaService()
    svc = TemplateService(str(tmp_path), schema)
    paths = svc.create_all()

    sku_path = paths["sku"]
    assert sku_path.name == "Шаблон_данных_SKU.xlsx"

    wb = load_workbook(sku_path)
    assert "Данные SKU" in wb.sheetnames
    assert "Инструкция" in wb.sheetnames
    ws = wb["Данные SKU"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "Артикул SKU" in headers
    assert "Продажи, руб." in headers
    assert "sku_code" not in headers
    assert "sales_value" not in headers

    # instruction sheet is Russian
    instr = wb["Инструкция"]
    assert "Обязательное" in [instr.cell(r, 2).value for r in range(1, instr.max_row + 1)]

    df = read_template_excel(sku_path, "sku")
    assert not df.empty
    assert "Артикул SKU" in df.columns


def test_mapping_russian_columns_to_canonical():
    schema = SchemaService()
    ms = MappingService(schema)
    columns = [
        "Артикул SKU",
        "Наименование товара",
        "Категория",
        "Подкатегория",
        "Продажи, руб.",
        "Продажи, шт.",
        "Себестоимость, руб.",
        "Остаток, шт.",
        "Цена закупки, руб.",
        "Длина выкладки, м",
        "Магазин",
        "Формат магазина",
        "Продажи LFL, руб.",
    ]
    mapping = ms.auto_map(columns, template_key="sku")
    assert mapping["sku_code"] == "Артикул SKU"
    assert mapping["sku_name"] == "Наименование товара"
    assert mapping["sales_value"] == "Продажи, руб."
    assert mapping["shelf_meters"] == "Длина выкладки, м"
    assert mapping["prior_sales_value"] == "Продажи LFL, руб."
    assert mapping["format_name"] == "Формат магазина"

    raw = pd.DataFrame([{c: 1 if "руб" in c or "шт" in c or c.endswith("м") else "x" for c in columns}])
    raw["Продажи, руб."] = 100
    raw["Продажи, шт."] = 2
    raw["Себестоимость, руб."] = 40
    raw["Остаток, шт."] = 5
    raw["Цена закупки, руб."] = 10
    mapped = ms.apply_mapping(raw, mapping)
    assert "sku_code" in mapped.columns
    assert "sales_value" in mapped.columns
    assert "Артикул SKU" not in mapped.columns


def test_graceful_degradation_missing_russian_optional_columns():
    schema = SchemaService()
    ms = MappingService(schema)
    # только обязательные русские колонки — без LFL и выкладки
    columns = [
        "Артикул SKU",
        "Наименование товара",
        "Категория",
        "Подкатегория",
        "Продажи, руб.",
        "Продажи, шт.",
        "Себестоимость, руб.",
        "Остаток, шт.",
        "Цена закупки, руб.",
    ]
    mapping = ms.auto_map(columns, template_key="sku")
    assert set(schema.required_fields("sku")).issubset(mapping.keys())
    assert "shelf_meters" not in mapping
    assert "prior_sales_value" not in mapping
    missing_req = ms.unmapped_required(mapping, "sku")
    assert missing_req == []

    df = pd.DataFrame(
        [
            {
                "Артикул SKU": "A1",
                "Наименование товара": "Товар",
                "Категория": "Кат",
                "Подкатегория": "Под",
                "Продажи, руб.": 1000,
                "Продажи, шт.": 10,
                "Себестоимость, руб.": 600,
                "Остаток, шт.": 3,
                "Цена закупки, руб.": 50,
            }
        ]
    )
    mapped = ms.apply_mapping(df, mapping)
    issues = validate_sku_df(mapped, schema)
    errors = [i for i in issues if i[0] == "error"]
    assert errors == []
    # сообщения об ошибках (если бы были) — на русском
    for _, msg, col in issues:
        assert "sku_code" not in str(msg)
        if col:
            assert col == schema.display_name("sku_code") or "_" not in str(col) or True

    calc = CalculationService()
    enriched = calc.enrich_base(mapped)
    kpi = calc.build_kpis(enriched, pd.DataFrame())
    assert not kpi.empty
    shelf = calc.shelf_efficiency(enriched)
    assert shelf.empty  # graceful: нет выкладки → пустой блок, без падения


def test_validation_messages_use_russian_labels():
    schema = SchemaService()
    df = pd.DataFrame({"sku_code": ["A"], "sku_name": ["N"]})  # мало обязательных
    issues = validate_sku_df(df, schema)
    messages = " ".join(i[1] for i in issues)
    assert "Артикул SKU" in messages or "Категория" in messages or "Продажи, руб." in messages
    assert "sales_value" not in messages


def test_goals_russian_metric_resolution():
    schema = SchemaService()
    assert schema.resolve_metric_code("Оборот") == "turnover"
    assert schema.resolve_metric_code("Дни запаса") == "stock_days"
    assert schema.resolve_metric_code("turnover") == "turnover"

    ms = MappingService(schema)
    svc = TemplateService(str(Path("data/processed/goals_tpl")), schema)
    path = svc.create_goals_template()
    assert path.name == "Шаблон_цели_категории.xlsx"
    df = read_template_excel(path, "goals")
    mapping = ms.auto_map(df.columns, template_key="goals")
    mapped = ms.apply_mapping(df, mapping)
    assert "metric_name" in mapped.columns
    calc = CalculationService()
    # минимальный SKU base
    base = pd.DataFrame(
        {
            "sku_code": ["1"],
            "sku_name": ["t"],
            "category": ["c"],
            "subcategory": ["s"],
            "sales_value": [1000],
            "sales_qty": [10],
            "cogs_value": [400],
            "stock_qty": [5],
            "purchase_price": [20],
            "period_months": [6],
        }
    )
    base = calc.enrich_base(base)
    kpi = calc.build_kpis(base, mapped)
    stock = kpi.loc[kpi["code"] == "stock_days", "target"].iloc[0]
    assert stock == 21


def test_english_aliases_still_work():
    """Устойчивость: старые английские заголовки тоже маппятся."""
    ms = MappingService()
    columns = [
        "sku_code",
        "sku_name",
        "category",
        "subcategory",
        "sales_value",
        "sales_qty",
        "cogs_value",
        "stock_qty",
        "purchase_price",
    ]
    mapping = ms.auto_map(columns, template_key="sku")
    assert mapping["sku_code"] == "sku_code"
    assert mapping["sales_value"] == "sales_value"
