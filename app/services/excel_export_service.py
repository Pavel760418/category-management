from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelExportService:
    def __init__(self):
        self.header_fill = PatternFill("solid", fgColor="1F4E78")
        self.header_font = Font(color="FFFFFF", bold=True, name="Calibri")
        self.title_font = Font(size=16, bold=True, name="Calibri", color="28251D")
        self.input_font = Font(color="0000FF", name="Calibri")
        self.formula_font = Font(color="000000", name="Calibri")
        self.link_font = Font(color="008000", name="Calibri")
        self.thin = Side(style="thin", color="D4D1CA")
        self.good = PatternFill("solid", fgColor="D9EAD3")
        self.warn = PatternFill("solid", fgColor="FFF2CC")
        self.bad = PatternFill("solid", fgColor="F4CCCC")

    def _style_table(self, ws, start_row, start_col, df, table_name):
        if df.empty:
            return
        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)

    def _write_df(self, ws, row, col, df, table_name=None):
        if df is None or df.empty:
            ws.cell(row=row, column=col, value="Нет доступных данных")
            return row + 2
        for j, c in enumerate(df.columns, start=col):
            cell = ws.cell(row=row, column=j, value=c)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, (_, rec) in enumerate(df.iterrows(), start=row + 1):
            for j, c in enumerate(df.columns, start=col):
                val = rec[c]
                ws.cell(row=i, column=j, value=None if pd.isna(val) else val)
        if table_name:
            self._style_table(ws, row, col, df, table_name)
        return row + len(df) + 3

    def export(self, output_path, base_df, kpi_df, abc_df, turnover_df, purchase_df, shelf_df, lfl_df, competitor_df, data_status_df, quality_df, limitations_df, instruction_text):
        wb = Workbook()
        ws = wb.active
        ws.title = "Защита категории"
        ws["B2"] = "🏆 Защита категории"
        ws["B2"].font = self.title_font
        row = 4
        row = self._write_df(ws, row, 2, kpi_df, "KPIStatus")
        row = self._write_df(ws, row, 2, data_status_df, "DataStatus")
        row = self._write_df(ws, row, 2, limitations_df, "Limitations")

        sheets = {
            "Данные SKU": base_df,
            "ABC-анализ": abc_df,
            "Оборачиваемость": turnover_df,
            "План закупок": purchase_df,
            "Выкладка": shelf_df,
            "LFL и цели": lfl_df,
            "Конкуренты": competitor_df,
            "Качество данных": quality_df,
            "Статус данных": data_status_df,
            "Ограничения анализа": limitations_df,
        }
        for name, df in sheets.items():
            wsx = wb.create_sheet(name)
            wsx["B2"] = name
            wsx["B2"].font = self.title_font
            self._write_df(wsx, 4, 2, df, f"T{abs(hash(name))%100000}")
            wsx.freeze_panes = "B5"
            wsx.column_dimensions["A"].width = 3
            for c in range(2, min(20, wsx.max_column + 1)):
                wsx.column_dimensions[get_column_letter(c)].width = 16

        wsi = wb.create_sheet("Инструкция")
        wsi["B2"] = "📋 Инструкция"
        wsi["B2"].font = self.title_font
        wsi["B4"] = instruction_text
        wsi["B4"].alignment = Alignment(wrap_text=True, vertical="top")
        wsi.row_dimensions[4].height = 600
        wsi.column_dimensions["B"].width = 120

        for wsx in wb.worksheets:
            wsx.column_dimensions["A"].width = 3
            for row in wsx.iter_rows():
                for cell in row:
                    cell.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
