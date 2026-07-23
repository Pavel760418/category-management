from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.schema_service import SchemaService, get_schema_service


class TemplateService:
    """Генерация русскоязычных Excel-шаблонов на основе metadata-слоя."""

    def __init__(self, template_dir: str, schema: SchemaService | None = None):
        self.template_dir = Path(template_dir)
        self.template_dir.mkdir(parents=True, exist_ok=True)
        self.schema = schema or get_schema_service()
        self.header_fill = PatternFill("solid", fgColor="1F4E78")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.note_fill = PatternFill("solid", fgColor="FFF2CC")
        self.required_fill = PatternFill("solid", fgColor="FCE4D6")
        self.optional_fill = PatternFill("solid", fgColor="E2EFDA")
        self.title_font = Font(size=14, bold=True, color="1F4E78")

    def _decorate_headers(self, ws, row, headers):
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=idx, value=h)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = max(16, min(28, len(str(h)) + 4))

    def _write_instruction_sheet(self, wb, template_key: str):
        tpl = self.schema.template(template_key)
        ws = wb.create_sheet(tpl["instruction_sheet_ru"])
        ws["A1"] = tpl["title_ru"]
        ws["A1"].font = self.title_font
        ws["A2"] = tpl.get("description_ru", "")
        ws["A4"] = "Инструкция по заполнению"
        ws["A4"].font = Font(bold=True)
        ws["A5"] = tpl.get("instruction_ru", "").strip()
        ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[5].height = 160
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 55
        ws.column_dimensions["D"].width = 28

        headers = ["Поле", "Обязательность", "Описание", "Пример"]
        start = 7
        self._decorate_headers(ws, start, headers)
        for i, row in enumerate(self.schema.field_guide_rows(template_key), start=start + 1):
            ws.cell(row=i, column=1, value=row["field"])
            req_cell = ws.cell(row=i, column=2, value=row["required"])
            req_cell.fill = self.required_fill if row["required"] == "Обязательное" else self.optional_fill
            ws.cell(row=i, column=3, value=row["description"])
            ws.cell(row=i, column=4, value=row["example"])

        ws["A3"] = (
            "Важно: названия колонок на листе данных должны совпадать с русскими заголовками шаблона. "
            "Внутренние технические ключи пользователю не нужны."
        )
        ws["A3"].fill = self.note_fill
        ws["A3"].alignment = Alignment(wrap_text=True)
        return ws

    def _create_from_meta(self, template_key: str) -> Path:
        tpl = self.schema.template(template_key)
        path = self.template_dir / tpl["file_name_ru"]
        wb = Workbook()
        ws = wb.active
        ws.title = tpl["sheet_name_ru"]

        fields = self.schema.template_fields(template_key, include_optional=True)
        headers = [self.schema.display_name(f) for f in fields]
        self._decorate_headers(ws, 1, headers)

        for r_idx, sample in enumerate(tpl.get("sample_rows") or [], start=2):
            for c_idx, canonical in enumerate(fields, start=1):
                ws.cell(row=r_idx, column=c_idx, value=sample.get(canonical))

        # Подсказка над таблицей через комментарий в строке заголовка не нужна —
        # отдельный лист «Инструкция» содержит полный гайд.
        if template_key == "sku" and "period_months" in fields:
            col_idx = fields.index("period_months") + 1
            from openpyxl.utils import get_column_letter

            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=1,
                formula2=12,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Период",
                error="Укажите целое число месяцев от 1 до 12",
            )
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}5000")

        self._write_instruction_sheet(wb, template_key)
        wb.save(path)
        return path

    def create_sku_template(self) -> Path:
        return self._create_from_meta("sku")

    def create_competitor_template(self) -> Path:
        return self._create_from_meta("competitor")

    def create_goals_template(self) -> Path:
        return self._create_from_meta("goals")

    def create_all(self) -> dict:
        return {
            "sku": self.create_sku_template(),
            "competitor": self.create_competitor_template(),
            "goals": self.create_goals_template(),
        }

    def list_templates_ru(self) -> list:
        """Список шаблонов для UI: русские имена файлов и подписи."""
        items = []
        for key in self.schema.template_keys():
            tpl = self.schema.template(key)
            items.append(
                {
                    "key": key,
                    "file_name": tpl["file_name_ru"],
                    "sheet_name": tpl["sheet_name_ru"],
                    "short_label": tpl["short_label_ru"],
                    "title": tpl["title_ru"],
                    "description": tpl["description_ru"],
                    "upload_label": tpl["upload_label_ru"],
                    "required_labels": [self.schema.display_name(f) for f in tpl["required_fields"]],
                    "optional_labels": [self.schema.display_name(f) for f in tpl.get("optional_fields") or []],
                }
            )
        return items
